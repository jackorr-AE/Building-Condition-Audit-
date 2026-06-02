"""Formatted Excel tables for Word copy-paste (Tables 4-1 and 4-2)."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Table 2-1 Building Condition Rating Index (report standard colours)
CONDITION_FILLS: dict[str, str] = {
    "excellent": "00B0F0",
    "good": "92D050",
    "fair": "FFFF00",
    "poor": "FFC000",
    "very poor": "FF0000",
}

HEADER_FILL = "005B2E"
HEADER_FONT_COLOR = "FFFFFF"
TITLE_FONT_COLOR = "005B2E"
DATA_FONT_COLOR = "000000"
NA_FILL = "FFFFFF"

THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

FONT_NAME = "Arial"


def condition_fill(value: str) -> str | None:
    key = (value or "").strip().lower()
    if not key or key in ("n/a", "not applicable", ""):
        return NA_FILL
    return CONDITION_FILLS.get(key)


def export_columns(fieldnames: list[str]) -> list[str]:
    return [h for h in fieldnames if (h or "").strip() and not h.endswith(" Comments")]


def load_condition_table(csv_path: Path, row_label_columns: set[str]) -> tuple[list[str], list[list[str]]]:
    with csv_path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError(f"No columns in {csv_path}")
        headers = export_columns(list(reader.fieldnames))
        rows: list[list[str]] = []
        for record in reader:
            rows.append([(record.get(h) or "").strip() for h in headers])
    return headers, rows


def _style_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: str,
    *,
    bold: bool = False,
    fill_hex: str | None = None,
    font_color: str = DATA_FONT_COLOR,
    align: str = "center",
    wrap: bool = False,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=FONT_NAME, size=10, bold=bold, color=font_color)
    cell.alignment = Alignment(
        horizontal=align,
        vertical="center",
        wrap_text=wrap,
    )
    cell.border = THIN_BORDER
    if fill_hex:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)


def write_formatted_table(
    ws: Worksheet,
    *,
    title: str,
    headers: list[str],
    rows: list[list[str]],
    row_label_columns: set[str],
) -> None:
    ncols = len(headers)
    if ncols == 0:
        raise ValueError("Table has no columns")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name=FONT_NAME, size=12, bold=True, color=TITLE_FONT_COLOR)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = 2
    for col, heading in enumerate(headers, start=1):
        _style_cell(
            ws,
            header_row,
            col,
            heading,
            bold=True,
            fill_hex=HEADER_FILL,
            font_color=HEADER_FONT_COLOR,
            align="center",
        )

    for r_idx, row_values in enumerate(rows, start=header_row + 1):
        for c_idx, (heading, value) in enumerate(zip(headers, row_values), start=1):
            is_label = heading in row_label_columns
            fill = None if is_label else condition_fill(value)
            if is_label:
                fill = NA_FILL
            _style_cell(
                ws,
                r_idx,
                c_idx,
                value,
                bold=False,
                fill_hex=fill,
                align="left" if is_label else "center",
            )

    # Column widths
    for col in range(1, ncols + 1):
        heading = headers[col - 1]
        letter = get_column_letter(col)
        if heading in row_label_columns:
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 14

    for row in range(header_row + 1, header_row + 1 + len(rows)):
        ws.row_dimensions[row].height = 18
    ws.row_dimensions[header_row].height = 22


def write_table_xlsx(
    csv_path: Path,
    xlsx_path: Path,
    *,
    title: str,
    row_label_columns: set[str],
    sheet_name: str,
) -> None:
    headers, rows = load_condition_table(csv_path, row_label_columns)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    write_formatted_table(
        ws,
        title=title,
        headers=headers,
        rows=rows,
        row_label_columns=row_label_columns,
    )
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def default_title_for_csv(csv_path: Path, table_number: str) -> str:
    stem = csv_path.stem
    if re.search(r"internal", stem, re.I):
        return f"Table {table_number} Overall Condition of Rooms and Components - Internal"
    if re.search(r"external", stem, re.I):
        return f"Table {table_number} Overall Condition of Rooms and Components - External"
    return f"Table {table_number}"
